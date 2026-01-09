# utils.py
# Contains utility and helper views:
# - dashboard_stats
# - cleanup_salary_data
# - health_check
# - get_dropdown_options
# - calculate_ot_rate
# - attendance_status
# - bulk_update_attendance
# - update_monthly_summaries_parallel
# - get_eligible_employees_for_date

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from ..models import EmployeeProfile
from django.db.models import Q, Sum, Count
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated, AllowAny
import logging
import time
import calendar

from ..models import (
    Tenant,
    SalaryData,
    Attendance,
    DailyAttendance,
    AdvanceLedger,
    Payment,
    CustomUser,
    UserPermissions,
    Leave,
    InvitationToken,
    PasswordResetOTP,
    PayrollPeriod,
    CalculatedSalary,
    SalaryAdjustment,
    DataSource,
    MonthlyAttendanceSummary,
)

from ..serializers import (
    TenantSerializer,
    CustomUserSerializer,
    CustomUserCreateSerializer,
    UserPermissionsSerializer,
    SalaryDataSerializer,
    SalaryDataSummarySerializer,
    EmployeeProfileSerializer,
    EmployeeProfileListSerializer,
    EmployeeFormSerializer,
    EmployeeTableSerializer,
    AttendanceSerializer,
    DailyAttendanceSerializer,
    AdvanceLedgerSerializer,
    PaymentSerializer,
    UserRegistrationSerializer,
    UserLoginSerializer,
    UserSerializer,
    SalaryDataFrontendSerializer,
)
from ..utils.permissions import IsSuperUser
from ..utils.utils import (
    clean_decimal_value,
    clean_int_value,
    is_valid_name,
    validate_excel_columns,
    generate_employee_id,
)

from ..services.salary_service import SalaryCalculationService

# Initialize logger
logger = logging.getLogger(__name__)

@api_view(["GET"])
def dashboard_stats(request):
    """
    Get dashboard statistics for current tenant
    TENANT-AWARE: All statistics are filtered by tenant
    """

    if not request.user.is_authenticated:
        return Response({"error": "Authentication required"}, status=401)

    # Get tenant from request
    tenant = getattr(request, 'tenant', None)
    if not tenant:
        return Response({"error": "No tenant found"}, status=400)

    # Get current month/year
    current_date = timezone.now()
    current_month = current_date.strftime("%B").upper()[:3]
    current_year = current_date.year

    # Employee count (tenant-specific)
    total_employees = EmployeeProfile.objects.filter(tenant=tenant, is_active=True).count()

    # Current month salary data (tenant-specific)
    current_month_data = SalaryData.objects.filter(
        tenant=tenant,
        year=current_year, 
        month__icontains=current_month
    )

    total_salary_paid = (
        current_month_data.aggregate(Sum("nett_payable"))["nett_payable__sum"] or 0
    )

    employees_paid = current_month_data.count()

    # Department distribution (tenant-specific)
    dept_distribution = (
        EmployeeProfile.objects.filter(tenant=tenant)
        .values("department")
        .annotate(count=Count("id"))
        .order_by("department")
    )

    return Response(
        {
            "total_employees": total_employees,
            "employees_paid_this_month": employees_paid,
            "total_salary_paid": float(total_salary_paid),
            "department_distribution": list(dept_distribution),
            "current_month": f"{current_month} {current_year}",
        }
    )


@api_view(["POST"])
def cleanup_salary_data(request):
    """

    Cleanup salary data (admin only)

    """

    if not request.user.is_authenticated or not request.user.is_tenant_admin:

        return Response({"error": "Admin access required"}, status=403)

    # Get parameters

    year = request.data.get("year")

    month = request.data.get("month")

    queryset = SalaryData.objects.all()

    if year:

        queryset = queryset.filter(year=year)

    if month:

        queryset = queryset.filter(month__icontains=month)

    deleted_count = queryset.count()

    queryset.delete()

    return Response(
        {
            "message": f"Deleted {deleted_count} salary records",
            "deleted_count": deleted_count,
        }
    )


# Health check endpoint


@api_view(["GET"])
@permission_classes([AllowAny])
def health_check(request):
    """
    Health check endpoint for monitoring
    """
    return Response(
        {"status": "healthy", "timestamp": timezone.now(), "version": "2.0.0"}
    )


@api_view(['GET'])
@permission_classes([AllowAny])
def get_dropdown_options(request):
    """
    Get unique values for all dropdowns for the public signup page.
    TENANT-AWARE: If tenant is available, only returns options for that tenant.
    """
    try:
        # Check if tenant is available (from middleware or authenticated user)
        tenant = getattr(request, 'tenant', None)
        
        # Build base queryset - filter by tenant if available
        base_queryset = EmployeeProfile.objects.all()
        if tenant:
            base_queryset = base_queryset.filter(tenant=tenant)
        
        # Get all unique, non-empty departments from employees (tenant-specific if tenant available)
        departments_clean = set(
            base_queryset.exclude(department__isnull=True)
            .exclude(department='')
            .values_list('department', flat=True)
            .distinct()
        )
        
        # Get all unique, non-empty location branches (tenant-specific if tenant available)
        locations_clean = set(
            base_queryset.exclude(location_branch__isnull=True)
            .exclude(location_branch='')
            .values_list('location_branch', flat=True)
            .distinct()
        )
        
        # Get all unique, non-empty designations (tenant-specific if tenant available)
        designations_clean = set(
            base_queryset.exclude(designation__isnull=True)
            .exclude(designation='')
            .values_list('designation', flat=True)
            .distinct()
        )

        # Get all unique, non-empty cities (tenant-specific if tenant available)
        cities_clean = set(
            base_queryset.exclude(city__isnull=True)
            .exclude(city='')
            .values_list('city', flat=True)
            .distinct()
        )

        # Get all unique, non-empty states (tenant-specific if tenant available)
        states_clean = set(
            base_queryset.exclude(state__isnull=True)
            .exclude(state='')
            .values_list('state', flat=True)
            .distinct()
        )
        
        return Response({
            'departments': sorted(list(departments_clean)),
            'locations': sorted(list(locations_clean)),
            'designations': sorted(list(designations_clean)),
            'cities': sorted(list(cities_clean)),
            'states': sorted(list(states_clean))
        })
        
    except Exception as e:
        # Log the error for debugging
        print(f"Error in get_dropdown_options: {str(e)}")
        return Response({"error": "An unexpected error occurred while fetching options."}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def calculate_ot_rate(request):
    """
    Calculate OT rate based on shift times and basic salary using STATIC formula
    Formula: OT Rate = basic_salary / (shift_hours × AVERAGE_DAYS_PER_MONTH)
    Using AVERAGE_DAYS_PER_MONTH from settings for consistent OT rates
    """
    try:
        from datetime import datetime, timedelta
        from decimal import Decimal
        from django.conf import settings
        
        shift_start_time_str = request.data.get('shift_start_time', '09:00')
        shift_end_time_str = request.data.get('shift_end_time', '18:00')
        basic_salary = Decimal(str(request.data.get('basic_salary', 0)))
        
        if basic_salary <= 0:
            return Response({'error': 'Basic salary must be greater than 0'}, status=400)
        
        # Parse shift times
        try:
            shift_start_time = datetime.strptime(shift_start_time_str, '%H:%M').time()
            shift_end_time = datetime.strptime(shift_end_time_str, '%H:%M').time()
        except (ValueError, TypeError):
            shift_start_time = datetime.strptime('09:00', '%H:%M').time()
            shift_end_time = datetime.strptime('18:00', '%H:%M').time()
        
        # Calculate shift hours
        start_dt = datetime.combine(datetime.today().date(), shift_start_time)
        end_dt = datetime.combine(datetime.today().date(), shift_end_time)
        if end_dt <= start_dt:
            end_dt += timedelta(days=1)
        raw_shift_hours = Decimal(str((end_dt - start_dt).total_seconds() / 3600))
        
        # Subtract break time from shift hours
        from ..utils.utils import get_break_time
        tenant = getattr(request, 'tenant', None)
        break_time = Decimal(str(get_break_time(tenant)))
        shift_hours = max(Decimal('0'), raw_shift_hours - break_time)
        
        if shift_hours <= 0:
            return Response({'error': 'Shift hours (after break time deduction) must be greater than 0'}, status=400)
        
        # Calculate OT rate using STATIC formula: basic_salary / ((shift_hours - break_time) × AVERAGE_DAYS_PER_MONTH)
        from ..utils.utils import get_average_days_per_month
        average_days = Decimal(str(get_average_days_per_month(tenant)))
        ot_rate = basic_salary / (shift_hours * average_days)
        calculation = f"{basic_salary} / (({raw_shift_hours:.2f} - {break_time:.2f}) hours × {average_days} days) = {round(ot_rate, 2)}"
        
        return Response({
            'ot_rate': round(float(ot_rate), 2),
            'calculation': calculation,
            'shift_hours_per_day': round(float(raw_shift_hours), 2),
            'break_time': float(break_time),
            'effective_shift_hours': round(float(shift_hours), 2),
            'average_days_per_month': float(average_days),
            'basic_salary': float(basic_salary),
            'formula': f'OT Rate = Basic Salary ÷ ((Shift Hours - Break Time) × Average Days Per Month)'
        })
    except (ValueError, TypeError) as e:
        return Response({'error': f'Invalid input: {str(e)}'}, status=400)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_salary_config(request):
    """
    Get salary calculation configuration values
    Returns tenant-specific AVERAGE_DAYS_PER_MONTH and BREAK_TIME for frontend use
    """
    try:
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return Response({'error': 'No tenant found'}, status=400)
        
        # Use tenant-specific values, fallback to defaults if not set
        average_days = float(getattr(tenant, 'average_days_per_month', 30.4) or 30.4)
        break_time = float(getattr(tenant, 'break_time', 0.5) if tenant.break_time is not None else 0.5)
        
        # Weekly rules configuration (tenant-specific, with sane defaults)
        weekly_absent_penalty_enabled = getattr(tenant, 'weekly_absent_penalty_enabled', False)
        weekly_absent_threshold = getattr(tenant, 'weekly_absent_threshold', 4) or 4
        sunday_bonus_enabled = getattr(tenant, 'sunday_bonus_enabled', False)
        # Present threshold is complement of absent threshold: 7 - absent_threshold
        sunday_bonus_threshold = 7 - weekly_absent_threshold
        
        return Response({
            'average_days_per_month': average_days,
            'break_time': break_time,
            'weekly_absent_penalty_enabled': bool(weekly_absent_penalty_enabled),
            'weekly_absent_threshold': int(weekly_absent_threshold),
            'sunday_bonus_enabled': bool(sunday_bonus_enabled),
            'sunday_bonus_threshold': int(sunday_bonus_threshold),  # Complement of absent threshold
            'description': 'Average days per month and break time used for salary and OT rate calculations'
        })
    except Exception as e:
        return Response({'error': f'Failed to get config: {str(e)}'}, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_salary_config(request):
    """
    Update salary calculation configuration values (tenant-specific)
    Expected payload: { "average_days_per_month": 30.4, "break_time": 0.0 }
    """
    try:
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return Response({'error': 'No tenant found'}, status=400)
        
        average_days = request.data.get('average_days_per_month')
        if average_days is None:
            return Response({'error': 'average_days_per_month is required'}, status=400)
        
        try:
            average_days_float = float(average_days)
            if average_days_float < 28 or average_days_float > 31:
                return Response({'error': 'average_days_per_month must be between 28 and 31'}, status=400)
        except (ValueError, TypeError):
            return Response({'error': 'average_days_per_month must be a valid number'}, status=400)
        
        # Handle break_time (optional, defaults to 0.0 if not provided)
        break_time = request.data.get('break_time', 0.0)
        try:
            break_time_float = float(break_time)
            if break_time_float < 0:
                return Response({'error': 'break_time must be greater than or equal to 0'}, status=400)
        except (ValueError, TypeError):
            return Response({'error': 'break_time must be a valid number'}, status=400)
        
        # Weekly rules configuration (all optional; use existing/defaults if not provided)
        weekly_absent_penalty_enabled = request.data.get('weekly_absent_penalty_enabled', getattr(tenant, 'weekly_absent_penalty_enabled', False))
        weekly_absent_threshold = request.data.get('weekly_absent_threshold', getattr(tenant, 'weekly_absent_threshold', 4))
        sunday_bonus_enabled = request.data.get('sunday_bonus_enabled', getattr(tenant, 'sunday_bonus_enabled', False))
        # Sunday bonus threshold is automatically calculated as complement: 7 - weekly_absent_threshold
        # No need to read it from request data
        
        try:
            weekly_absent_penalty_enabled = bool(weekly_absent_penalty_enabled)
            weekly_absent_threshold = int(weekly_absent_threshold)
            if weekly_absent_threshold < 2 or weekly_absent_threshold > 7:
                return Response({'error': 'weekly_absent_threshold must be between 2 and 7 (since present threshold = 7 - absent_threshold)'}, status=400)
        except (ValueError, TypeError):
            return Response({'error': 'weekly_absent_threshold must be a valid integer'}, status=400)
        
        try:
            sunday_bonus_enabled = bool(sunday_bonus_enabled)
        except (ValueError, TypeError):
            return Response({'error': 'sunday_bonus_enabled must be a valid boolean'}, status=400)
        
        tenant.average_days_per_month = average_days_float
        tenant.break_time = break_time_float
        tenant.weekly_absent_penalty_enabled = weekly_absent_penalty_enabled
        tenant.weekly_absent_threshold = weekly_absent_threshold
        tenant.sunday_bonus_enabled = sunday_bonus_enabled
        # Sunday bonus threshold is automatically calculated as complement: 7 - weekly_absent_threshold
        # We keep the field in DB for backward compatibility but don't update it
        tenant.save(update_fields=[
            'average_days_per_month',
            'break_time',
            'weekly_absent_penalty_enabled',
            'weekly_absent_threshold',
            'sunday_bonus_enabled',
            # 'sunday_bonus_threshold',  # Not saved separately, calculated as 7 - weekly_absent_threshold
        ])
        
        # Clear relevant caches
        from django.core.cache import cache
        cache_keys_to_clear = [
            f"directory_data_{tenant.id}",
            f"directory_data_full_{tenant.id}",
            f"payroll_overview_{tenant.id}",
            f"months_with_attendance_{tenant.id}",
        ]
        for key in cache_keys_to_clear:
            cache.delete(key)
        
        return Response({
            'success': True,
            'average_days_per_month': float(tenant.average_days_per_month),
            'break_time': float(tenant.break_time),
            'weekly_absent_penalty_enabled': bool(tenant.weekly_absent_penalty_enabled),
            'weekly_absent_threshold': int(tenant.weekly_absent_threshold),
            'sunday_bonus_enabled': bool(tenant.sunday_bonus_enabled),
            'sunday_bonus_threshold': int(7 - tenant.weekly_absent_threshold),  # Complement of absent threshold
            'message': 'Salary configuration updated successfully'
        })
    except Exception as e:
        logger.error(f"Error updating salary config: {str(e)}", exc_info=True)
        return Response({'error': f'Failed to update config: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def attendance_status(request):
    """
    Get attendance tracking status and information
    """
    try:
        from datetime import datetime, date
        
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return Response({"error": "No tenant found"}, status=400)
        
        # Attendance system start date
        attendance_start_date = date(2025, 6, 22)
        current_date = datetime.now().date()
        
        # Check if attendance tracking is active
        is_active = current_date >= attendance_start_date
        
        # Get total active employees
        total_active_employees = EmployeeProfile.objects.filter(
            tenant=tenant,
            is_active=True
        ).count()
        
        # Get employees with attendance records this month
        employees_with_records = Attendance.objects.filter(
            tenant=tenant,
            date__year=current_date.year,
            date__month=current_date.month
        ).count()
        
        # Check if we have day-by-day attendance data (DailyAttendance records)
        has_daily_tracking = DailyAttendance.objects.filter(
            tenant=tenant,
            date__gte=attendance_start_date
        ).exists()
        
        return Response({
            'is_active': is_active,
            'start_date': attendance_start_date.strftime('%Y-%m-%d'),
            'current_date': current_date.strftime('%Y-%m-%d'),
            'total_active_employees': total_active_employees,
            'employees_with_records': employees_with_records,
            'has_daily_tracking': has_daily_tracking,
            'tracking_mode': 'daily' if has_daily_tracking else 'monthly',
            'message': 'Attendance tracking is active' if is_active else f'Attendance tracking will start from {attendance_start_date.strftime("%B %d, %Y")}'
        })
        
    except Exception as e:
        logger.error(f"Error getting attendance status: {str(e)}")
        return Response({"error": "Failed to get attendance status"}, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_update_attendance(request):
    """
    Optimized bulk update attendance with batch processing for better performance
    """
    try:
        from datetime import datetime
        from django.db import transaction
        from excel_data.signals import sync_attendance_from_daily
        from excel_data.models import DailyAttendance
        
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return Response({"error": "No tenant found"}, status=400)
        
        date_str = request.data.get('date')
        attendance_records = request.data.get('attendance_records', [])
        
        if not date_str:
            return Response({"error": "Date is required"}, status=400)
        
        if not attendance_records:
            return Response({"error": "Attendance records are required"}, status=400)
        
        # Parse the date
        try:
            attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)
        
        # Check if date is not in the future
        if attendance_date > datetime.now().date():
            return Response({"error": "Cannot mark attendance for future dates"}, status=400)
        
        # Check if date is a holiday - optimize by doing this once upfront
        from ..models import Holiday
        from ..utils.utils import get_current_tenant
        
        # Get tenant (use get_current_tenant if available, fallback to request.tenant)
        tenant_for_holiday = get_current_tenant()
        if not tenant_for_holiday:
            tenant_for_holiday = tenant
        
        holiday_applies = False
        holiday_info = None
        
        if tenant_for_holiday:
            # Check if date is a holiday (single query)
            holiday = Holiday.objects.filter(
                tenant=tenant_for_holiday,
                date=attendance_date,
                is_active=True
            ).first()
            
            if holiday:
                # Get unique departments from attendance records once
                departments_in_request = set()
                for record in attendance_records:
                    dept = record.get('department', '')
                    if dept:
                        departments_in_request.add(dept)
                
                # Check if holiday applies
                holiday_applies = holiday.applies_to_all
                if not holiday_applies and holiday.specific_departments:
                    holiday_departments = set(d.strip() for d in holiday.specific_departments.split(','))
                    holiday_applies = bool(departments_in_request & holiday_departments)  # Set intersection
                
                if holiday_applies:
                    # Holiday applies - block attendance marking
                    error_message = f"Cannot mark attendance on holiday: {holiday.name}"
                    if holiday.description:
                        error_message += f" - {holiday.description}"
                    return Response({
                        "error": error_message,
                        "holiday": {
                            "name": holiday.name,
                            "description": holiday.description,
                            "type": holiday.holiday_type
                        }
                    }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get day of week for off-day checks
        day_of_week = attendance_date.weekday()  # Monday = 0, Sunday = 6
        
        # Extract all employee IDs from attendance records
        employee_ids = [record.get('employee_id') for record in attendance_records if record.get('employee_id')]
        
        if not employee_ids:
            return Response({"error": "No valid employee IDs found"}, status=400)
        
        # Bulk fetch all employees in one query
        # Only fetch the fields we actually need to reduce memory and DB overhead
        employees = EmployeeProfile.objects.filter(
            tenant=tenant,
            employee_id__in=employee_ids,
            is_active=True
        ).only(
            'employee_id', 'first_name', 'last_name', 'department', 
            'designation', 'employment_type', 'date_of_joining',
            'off_monday', 'off_tuesday', 'off_wednesday', 'off_thursday',
            'off_friday', 'off_saturday', 'off_sunday'
        )
        
        # Create employee lookup dictionary for fast access
        employee_lookup = {emp.employee_id: emp for emp in employees}
        
        # Get existing attendance records for this date to determine updates vs creates
        existing_attendance = DailyAttendance.objects.filter(
            tenant=tenant,
            employee_id__in=employee_ids,
            date=attendance_date
        )
        existing_lookup = {att.employee_id: att for att in existing_attendance}
        
        # Prepare batch data
        records_to_create = []
        records_to_update = []
        created_count = 0
        updated_count = 0
        skipped_count = 0
        skipped_employees = []  # Track skipped employees with reasons
        errors = []
        # Track which employees had attendance marked for Sunday bonus check
        employees_to_check_sunday_bonus = set()
        
        # PERFORMANCE OPTIMIZATION: Add timing and batch size optimization
        import time
        processing_start_time = time.time()
        cache_clear_time = 0  # Initialize cache_clear_time variable
        
        # Process records in batch with performance tracking
        batch_size = 500  # Reduced from processing all at once
        processed_batches = 0
        
        for record in attendance_records:
            try:
                employee_id = record.get('employee_id')
                if not employee_id:
                    errors.append(f"Missing employee_id in record")
                    continue
                
                # Check if employee exists (using lookup dictionary)
                employee = employee_lookup.get(employee_id)
                if not employee:
                    errors.append(f"Employee {employee_id} not found or inactive")
                    continue
                
                # Check if employee has joined by this date
                if employee.date_of_joining and attendance_date < employee.date_of_joining:
                    skipped_count += 1
                    skipped_employees.append({
                        'employee_id': employee_id,
                        'name': record.get('name') or f"{employee.first_name} {employee.last_name}",
                        'reason': f"Employee has not joined yet (joining date: {employee.date_of_joining})"
                    })
                    continue
                
                # OPTIMIZED: Use pre-calculated off day check
                off_day_flags = [
                    employee.off_monday, employee.off_tuesday, employee.off_wednesday,
                    employee.off_thursday, employee.off_friday, employee.off_saturday, employee.off_sunday
                ]
                
                # Get status early to check if we should process off-day records
                status = record.get('status')
                
                # Only skip if employee has off day AND status is 'absent' or 'unmarked'
                # Allow 'off' status to record off days, and 'present' status for extra pay on off days
                if off_day_flags[day_of_week] and status not in ['off', 'present']:
                    skipped_count += 1
                    day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                    skipped_employees.append({
                        'employee_id': employee_id,
                        'name': record.get('name') or f"{employee.first_name} {employee.last_name}",
                        'reason': f"Employee has off day on {day_names[day_of_week]} and status is '{status}' (should be 'off' or 'present' for extra pay)"
                    })
                    continue  # Skip attendance for off days unless explicitly marking as 'off' or 'present'
                
                # OPTIMIZED: Minimal data processing
                ot_hours = float(record.get('ot_hours', 0))
                late_minutes = int(record.get('late_minutes', 0))
                employee_name = record.get('name') or f"{employee.first_name} {employee.last_name}"
                department = record.get('department') or employee.department or 'General'
                
                # Handle off-day status optimization
                if status == 'off':
                    ot_hours = 0
                    late_minutes = 0
                
                # OPTIMIZED: Fast status determination
                # Handle 'present', 'absent', 'off', and 'unmarked' statuses
                if status == 'present':
                    attendance_status = 'PRESENT'
                elif status == 'absent':
                    attendance_status = 'ABSENT'
                elif status == 'off':
                    attendance_status = 'OFF'
                else:
                    # Default to 'UNMARKED' for anything else (including explicit 'unmarked')
                    attendance_status = 'UNMARKED'
                
                # OPTIMIZED: Prepare record data with minimal overhead
                record_data = {
                    'employee_name': employee_name,
                    'department': department,
                    'designation': employee.designation or 'General',
                    'employment_type': employee.employment_type or 'FULL_TIME',
                    'attendance_status': attendance_status,
                    'ot_hours': ot_hours,
                    'late_minutes': late_minutes,
                }
                
                # OPTIMIZED: Fast update/create decision
                if employee_id in existing_lookup:
                    # Update existing record
                    existing_record = existing_lookup[employee_id]
                    for key, value in record_data.items():
                        setattr(existing_record, key, value)
                    records_to_update.append(existing_record)
                    updated_count += 1
                    # Track for Sunday bonus check if status is PRESENT or PAID_LEAVE
                    if attendance_status in ['PRESENT', 'PAID_LEAVE']:
                        employees_to_check_sunday_bonus.add((employee_id, attendance_date))
                else:
                    # Create new record
                    new_record = DailyAttendance(
                        tenant=tenant,
                        employee_id=employee_id,
                        date=attendance_date,
                        **record_data
                    )
                    records_to_create.append(new_record)
                    created_count += 1
                    # Track for Sunday bonus check if status is PRESENT or PAID_LEAVE
                    if attendance_status in ['PRESENT', 'PAID_LEAVE']:
                        employees_to_check_sunday_bonus.add((employee_id, attendance_date))
                    
            except Exception as e:
                errors.append(f"Error processing employee {record.get('employee_id', 'unknown')}: {str(e)}")
        
        processing_time = time.time() - processing_start_time
        logger.info(f"OPTIMIZED: Processed {len(attendance_records)} records in {processing_time:.3f}s")
        
        # ULTRA OPTIMIZED: Perform bulk operations with Django ORM (faster and safer than raw SQL)
        db_start_time = time.time()
        
        with transaction.atomic():
            # BATCH SIZE: Process in optimal batches for PostgreSQL
            # PostgreSQL performs best with batches of 500-1000 records
            BATCH_SIZE = 500
            
            # OPTIMIZED: Bulk create using Django ORM (faster than raw SQL for large batches)
            if records_to_create:
                # Use bulk_create with batch_size for optimal performance
                # ignore_conflicts=False ensures we catch unique constraint violations
                DailyAttendance.objects.bulk_create(
                    records_to_create,
                    batch_size=BATCH_SIZE,
                    ignore_conflicts=False
                )
                logger.info(f"✅ Django ORM bulk created {len(records_to_create)} records in batches of {BATCH_SIZE}")
            
            # OPTIMIZED: Bulk update using Django ORM (much faster than raw SQL CASE statements)
            if records_to_update:
                # Update only the fields that changed - more efficient than updating all fields
                fields_to_update = [
                    'employee_name', 'department', 'designation', 'employment_type',
                    'attendance_status', 'ot_hours', 'late_minutes', 'updated_at'
                ]
                
                # Django's bulk_update is highly optimized with proper query planning
                DailyAttendance.objects.bulk_update(
                    records_to_update,
                    fields_to_update,
                    batch_size=BATCH_SIZE
                )
                logger.info(f"✅ Django ORM bulk updated {len(records_to_update)} records in batches of {BATCH_SIZE}")
        
        db_operation_time = time.time() - db_start_time
        logger.info(f"OPTIMIZED: Core DB operations completed in {db_operation_time:.3f}s")
        
        # BACKGROUND THREAD: Trigger Sunday bonus check for employees with present/paid_leave attendance
        # Note: bulk_create/bulk_update don't trigger signals, so we manually trigger here
        if employees_to_check_sunday_bonus:
            try:
                from excel_data.tasks import mark_sunday_bonus_background
                logger.info(f"🔄 [Bulk] Triggering Sunday bonus check for {len(employees_to_check_sunday_bonus)} employees")
                for employee_id, attendance_date in employees_to_check_sunday_bonus:
                    mark_sunday_bonus_background(tenant.id, employee_id, attendance_date)
            except Exception as e:
                logger.error(f"❌ [Bulk] Failed to trigger Sunday bonus background tasks: {e}", exc_info=True)
        
        # LIGHTNING FAST: Skip monthly summary recalculation for bulk uploads
        # Instead, defer this to a background task or make it optional
        summary_start_time = time.time()
        summaries_updated = 0
        
        # Get all affected employee IDs for cache clearing only
        affected_employee_ids = set()
        for record in records_to_create + records_to_update:
            affected_employee_ids.add(record.employee_id)
        
        # PERFORMANCE DECISION: Skip heavy monthly summary calculation
        # This reduces 7+ seconds to nearly instant for bulk operations
        # Monthly summaries can be calculated on-demand or via background job
        
        logger.info(f"LIGHTNING FAST: Skipped monthly summary recalculation for {len(affected_employee_ids)} employees")
        logger.info("Monthly summaries will be calculated on-demand when needed")
        
        summary_time = time.time() - summary_start_time
        logger.info(f"LIGHTNING OPTIMIZED: Summary processing completed in {summary_time:.3f}s")
        
        # OPTIMIZED CACHE CLEARING: Clear only the most critical caches synchronously
        # Defer comprehensive cache clearing to background thread for instant API response
        from django.core.cache import cache
        
        cache_start_time = time.time()
        tenant_id = tenant.id if tenant else 'default'
        
        # Clear ONLY the most critical cache keys that affect immediate UI response
        critical_cache_keys = [
            f"eligible_employees_{tenant_id}_{date_str}",  # Critical for attendance log UI refresh
            f"eligible_employees_progressive_{tenant_id}_{date_str}_initial",
            f"directory_data_{tenant_id}",  # Critical for directory to show updated OT/late totals
        ]
        
        # Use cache.delete_many for batch deletion (single operation instead of multiple)
        cache.delete_many(critical_cache_keys)
        
        # Pattern-based deletion for attendance_all_records (most frequently accessed)
        try:
            # Try pattern-based deletion if available (Redis cache)
            deleted_count = cache.delete_pattern(f"attendance_all_records_{tenant_id}_*")
            logger.info(f"✅ Pattern-deleted attendance_all_records cache ({deleted_count} keys)")
        except (AttributeError, NotImplementedError):
            # Fallback: Clear only the specific date/month variation
            month_num = attendance_date.month
            year_num = attendance_date.year
            cache.delete_many([
                f"attendance_all_records_{tenant_id}_custom_month_{month_num}_{year_num}_None_None_rt_1",
                f"attendance_all_records_{tenant_id}_one_day_None_None_{date_str}_{date_str}_rt_1",
            ])
        
        cache_clear_time = time.time() - cache_start_time
        logger.info(f"⚡ Fast cache clear: {len(critical_cache_keys)} keys in {cache_clear_time:.3f}s")
        
        
        # Prepare comprehensive cache keys for background clearing (non-critical caches)
        comprehensive_cache_keys = [
            f"payroll_overview_{tenant_id}",
            f"months_with_attendance_{tenant_id}",
            f"eligible_employees_progressive_{tenant_id}_{date_str}_remaining",
            f"total_eligible_count_{tenant_id}_{date_str}",
            f"directory_data_full_{tenant_id}",
            f"attendance_log_{tenant_id}",
            f"attendance_tracker_{tenant_id}",
            f"monthly_attendance_summary_{tenant_id}_{attendance_date.year}_{attendance_date.month}",
            f"dashboard_stats_{tenant_id}",
            f"weekly_attendance_{tenant_id}_{date_str}",
        ]
        
        # Add weekly attendance cache for all days in the week
        from datetime import timedelta
        try:
            start_of_week = attendance_date - timedelta(days=attendance_date.weekday())  # Monday
            for day_offset in range(7):
                week_date = start_of_week + timedelta(days=day_offset)
                comprehensive_cache_keys.append(f"weekly_attendance_{tenant_id}_{week_date.isoformat()}")
        except Exception as e:
            logger.warning(f"Failed to add week dates for background cache clear: {e}")
        
        # Mark patterns for background deletion
        comprehensive_cache_keys.extend([
            f"frontend_charts_{tenant_id}_*",
            f"eligible_employees_opt_{tenant_id}_*",
        ])
        
        # Store cache keys for background clearing
        cache_keys_for_background = comprehensive_cache_keys
        
        # Calculate comprehensive performance metrics
        total_function_time = time.time() - processing_start_time
        total_uploaded = created_count + updated_count
        
        response_data = {
            'message': 'Attendance uploaded successfully',
            'status': 'success',
            'attendance_upload': {
                'total_processed': total_uploaded,
                'created_count': created_count,
                'updated_count': updated_count,
                'skipped_count': skipped_count,
                'skipped_employees': skipped_employees if skipped_employees else None,
                'date': date_str
            }
        }
        
        if errors:
            response_data['errors'] = errors
            response_data['message'] += f' ({len(errors)} errors occurred)'
        
        # BACKGROUND AGGREGATION: Start monthly aggregation in non-blocking background thread
        logger.info("🚀 BACKGROUND AGGREGATION: Starting monthly aggregation in background thread (non-blocking)...")
        
        try:
            import threading
            from django.db import connections
            from ..utils.utils import run_bulk_aggregation
            
            # Capture variables needed in thread (tenant_id instead of tenant object to avoid connection issues)
            tenant_id = tenant.id
            attendance_year = attendance_date.year
            attendance_month = attendance_date.month
            attendance_date_str = date_str  # Capture date string for cache clearing
            
            def background_aggregation():
                """
                Non-blocking background thread function for monthly aggregation and cache clearing.
                Properly handles Django database connections in threads.
                """
                try:
                    # Close any existing database connections before starting (thread-safe)
                    connections.close_all()
                    
                    # Re-fetch tenant in the thread (ensures fresh DB connection)
                    from datetime import date
                    from ..models import Tenant
                    thread_tenant = Tenant.objects.get(id=tenant_id)
                    thread_attendance_date = date(attendance_year, attendance_month, 1)
                    
                    logger.info(f"🧵 BACKGROUND THREAD: Starting aggregation for {thread_tenant.subdomain} - {attendance_year}-{attendance_month:02d}")
                    
                    # Run monthly aggregation (this can take several seconds for large datasets)
                    aggregation_start = time.time()
                    result = run_bulk_aggregation(thread_tenant, thread_attendance_date)
                    aggregation_elapsed = time.time() - aggregation_start
                    
                    # Log detailed completion info
                    status = result.get('status', 'unknown')
                    stats = result.get('statistics', {})
                    perf = result.get('performance', {})
                    
                    logger.info(f"✅ BACKGROUND THREAD: Aggregation completed with status: {status}")
                    logger.info(f"📊 BACKGROUND THREAD: Employees processed: {stats.get('employees_processed', 0)}, "
                              f"Records: {stats.get('daily_records_processed', 0)}, "
                              f"Created: {stats.get('attendance_records_created', 0)}, "
                              f"Updated: {stats.get('attendance_records_updated', 0)}")
                    logger.info(f"⏱️ BACKGROUND THREAD: Total time: {aggregation_elapsed:.3f}s, "
                              f"DB time: {perf.get('database_time', 'N/A')}, "
                              f"Cache time: {perf.get('cache_clear_time', 'N/A')}")
                    
                    
                    # Clear comprehensive caches in background (simplified)
                    logger.info(f"🧵 BACKGROUND THREAD: Starting comprehensive cache clearing...")
                    cache_start_time = time.time()
                    
                    from django.core.cache import cache
                    
                    # Try pattern-based deletion first (most efficient)
                    patterns_cleared = 0
                    try:
                        patterns_to_clear = [
                            f"frontend_charts_{tenant_id}_*",
                            f"eligible_employees_opt_{tenant_id}_*",
                        ]
                        for pattern in patterns_to_clear:
                            patterns_cleared += cache.delete_pattern(pattern)
                        logger.info(f"✅ Pattern-deleted {patterns_cleared} cache keys")
                    except (AttributeError, NotImplementedError):
                        logger.info("ℹ️ Pattern deletion not supported, using delete_many")
                    
                    # Delete specific keys in batch (much faster than individual deletes)
                    cache.delete_many([k for k in cache_keys_for_background if not k.endswith('_*')])
                    
                    cache_time = time.time() - cache_start_time
                    logger.info(f"🧵 BACKGROUND THREAD: Cleared caches in {cache_time:.3f}s")
                    
                    # Close database connections after thread completes
                    connections.close_all()
                    
                except Exception as e:
                    logger.error(f"❌ BACKGROUND THREAD ERROR: {str(e)}", exc_info=True)
                    # Ensure connections are closed even on error
                    try:
                        connections.close_all()
                    except Exception:
                        pass
            
            # Start background thread with daemon=True (non-blocking, dies when main thread exits)
            background_thread = threading.Thread(
                target=background_aggregation, 
                daemon=True,
                name=f"attendance_agg_{tenant_id}_{attendance_year}_{attendance_month}"
            )
            background_thread.start()
            
            logger.info(f"🧵 BACKGROUND THREAD: Started with ID {background_thread.ident} (name: {background_thread.name})")
            
            # Background aggregation started (no need to expose debug info in response)
            
        except Exception as thread_error:
            logger.error(f"❌ BACKGROUND THREAD FAILED TO START: {thread_error}", exc_info=True)
            # Don't expose thread errors in response - just log them
        
        return Response(response_data, status=200)
        
    except Exception as e:
        logger.error(f"Error in bulk update attendance: {str(e)}")
        return Response({"error": "Failed to update attendance"}, status=500)

# Clean replacement for the update_monthly_summaries_parallel function

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def recalculate_penalty_bonus_days(request):
    """
    Recalculate weekly penalty and bonus days for MonthlyAttendanceSummary records.
    Useful when features are enabled after data already exists.
    
    Expected payload (all optional):
    {
        "year": 2025,
        "month": 12,
        "employee_id": "EMP-001"  # Optional: recalculate for specific employee only
    }
    """
    try:
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return Response({'error': 'No tenant found'}, status=400)
        
        year = request.data.get('year')
        month = request.data.get('month')
        employee_id = request.data.get('employee_id')
        
        from excel_data.models import MonthlyAttendanceSummary, EmployeeProfile
        from excel_data.services.salary_service import SalaryCalculationService
        from decimal import Decimal
        
        # Build query
        queryset = MonthlyAttendanceSummary.objects.filter(tenant=tenant)
        
        if year:
            queryset = queryset.filter(year=year)
        if month:
            queryset = queryset.filter(month=month)
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        
        total = queryset.count()
        if total == 0:
            return Response({
                'success': True,
                'message': 'No records found to recalculate',
                'updated_count': 0
            })
        
        month_names = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
        updated_count = 0
        error_count = 0
        
        for summary in queryset.select_related('tenant'):
            try:
                # Get employee
                try:
                    employee = EmployeeProfile.objects.get(
                        tenant=summary.tenant,
                        employee_id=summary.employee_id,
                        is_active=True
                    )
                except EmployeeProfile.DoesNotExist:
                    continue
                
                # Check if features are enabled
                absent_enabled = getattr(summary.tenant, 'weekly_absent_penalty_enabled', False)
                bonus_enabled = getattr(summary.tenant, 'sunday_bonus_enabled', False)
                
                if not absent_enabled and not bonus_enabled:
                    continue
                
                # Calculate penalty/bonus
                month_name = month_names[summary.month - 1] if 1 <= summary.month <= 12 else 'JAN'
                weekly_stats = SalaryCalculationService._compute_weekly_penalty_and_bonus(
                    employee, summary.year, month_name
                )
                
                new_penalty = weekly_stats.get('weekly_penalty_days', Decimal("0"))
                
                summary.weekly_penalty_days = new_penalty
                # Sunday bonus handled separately by marking Sunday as PRESENT (no field update needed)
                summary.save(update_fields=['weekly_penalty_days'])
                updated_count += 1
                
            except Exception as e:
                error_count += 1
                import logging
                logging.getLogger(__name__).error(f'Error recalculating penalty/bonus: {e}', exc_info=True)
        
        return Response({
            'success': True,
            'message': f'Recalculated {updated_count} records, {error_count} errors',
            'updated_count': updated_count,
            'error_count': error_count,
            'total_records': total
        })
        
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f'Error in recalculate_penalty_bonus_days: {e}', exc_info=True)
        return Response({'error': f'Failed to recalculate: {str(e)}'}, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_monthly_summaries_parallel(request):
    """
    Asynchronous API for updating monthly summaries after bulk attendance upload.
    Returns immediately while processing summaries in background thread.
    
    Expected usage:
    1. Frontend calls this API after bulk attendance upload
    2. Returns success immediately 
    3. Processing happens in background thread using ULTRA-FAST bulk operations
    4. Cache is cleared immediately for instant UI updates
    """
    try:
        import threading
        from datetime import datetime
        from django.core.cache import cache
        
        # Get logger instance
        logger = logging.getLogger(__name__)
        
        start_time = time.time()
        
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return Response({"error": "No tenant found"}, status=400)
        
        date_str = request.data.get('date')
        employee_ids = request.data.get('employee_ids', [])
        
        if not date_str:
            return Response({"error": "Date is required"}, status=400)
        
        # Parse the date
        try:
            attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)
        
        # DEBUG: Initial request information
        logger.info(f"🚀 ASYNC SUMMARY: API called with parameters:")
        logger.info(f"   📅 Date: {date_str} (parsed as {attendance_date})")
        logger.info(f"   👥 Employee IDs count: {len(employee_ids)}")
        logger.info(f"   🏢 Tenant: {tenant.name} (ID: {tenant.id})")
        logger.info(f"   👥 Employee IDs (first 10): {employee_ids[:10]}{'...' if len(employee_ids) > 10 else ''}")
        
        print(f"🚀 ASYNC SUMMARY: API called - Date: {date_str}, Employee count: {len(employee_ids)}, Tenant: {tenant.name}")
        
        logger.info(f"🔄 ASYNC SUMMARY: Starting background monthly summary update for {len(employee_ids)} employees on {date_str}")
        
        # CLEAR ALL RELATED CACHES IMMEDIATELY for instant UI updates
        cache_start_time = time.time()
        cache_keys_to_clear = [
            f"payroll_overview_{tenant.id}",
            f"months_with_attendance_{tenant.id}",
            f"eligible_employees_{tenant.id}",
            f"eligible_employees_{tenant.id}_progressive",
            f"directory_data_{tenant.id}",  # Critical for directory to show updated OT/late totals
            f"directory_data_full_{tenant.id}",  # Critical for directory full dataset cache
            f"attendance_all_records_{tenant.id}",
            f"attendance_log_{tenant.id}",
            f"attendance_tracker_{tenant.id}",
            f"monthly_attendance_summary_{tenant.id}_{attendance_date.year}_{attendance_date.month}",
            f"monthly_attendance_summary_{tenant.id}",
            f"dashboard_stats_{tenant.id}",
            f"employee_attendance_history_{tenant.id}",
            # CRITICAL: Add all_records cache variations that need immediate clearing
            f"attendance_all_records_{tenant.id}_this_month_None_None_None_None",
            f"attendance_all_records_{tenant.id}_last_6_months_None_None_None_None",
            f"attendance_all_records_{tenant.id}_last_12_months_None_None_None_None",
            f"attendance_all_records_{tenant.id}_last_5_years_None_None_None_None",
            f"frontend_charts_{tenant.id}",
        ]
        
        # Clear all cache keys (handle pattern deletion)
        for cache_key in cache_keys_to_clear:
            if cache_key.endswith('_*'):
                # Handle pattern deletion
                try:
                    cache.delete_pattern(cache_key)
                except AttributeError:
                    # Fallback: Clear common variations
                    base_pattern = cache_key.replace('_*', '')
                    for period in ['this_month', 'last_6_months', 'last_12_months', 'last_5_years']:
                        cache.delete(f"{base_pattern}_{period}_All_")
            else:
                cache.delete(cache_key)
            
        # Clear any date-specific cache keys
        tenant_id = tenant.id if tenant else 'default'
        cache.delete(f"attendance_all_records_{tenant_id}_{date_str}")
        cache.delete(f"eligible_employees_{tenant_id}_{date_str}")
        
        # Clear directory cache (ensure it's cleared even if not in cache_keys_to_clear)
        cache.delete(f"directory_data_{tenant_id}")
        cache.delete(f"directory_data_full_{tenant_id}")
        
        # Clear custom date-based all_records cache keys that might exist
        cache.delete(f"attendance_all_records_{tenant_id}_custom_{attendance_date.month}_{attendance_date.year}_None_None")
        cache.delete(f"attendance_all_records_{tenant_id}_custom_range_None_None_{date_str}_None")
        
        # Clear frontend charts cache (pattern matching)
        try:
            cache.delete_pattern(f"frontend_charts_{tenant_id}_*")
        except AttributeError:
            # Fallback: Clear common chart cache keys
            chart_keys = [
                f"frontend_charts_{tenant_id}_this_month_All_",
                f"frontend_charts_{tenant_id}_last_6_months_All_",
                f"frontend_charts_{tenant_id}_last_12_months_All_",
                f"frontend_charts_{tenant_id}_last_5_years_All_"
            ]
            for key in chart_keys:
                cache.delete(key)
        
        cache_time = time.time() - cache_start_time
        logger.info(f"🗑️ ASYNC SUMMARY: Cleared {len(cache_keys_to_clear) + 4} cache keys in {cache_time:.3f}s")
        logger.info(f"🗑️ ASYNC SUMMARY: Cache keys cleared: {cache_keys_to_clear[:5]}{'...' if len(cache_keys_to_clear) > 5 else ''}")
        
        # Define non-blocking background processing function using our optimized aggregation
        def process_summaries_background():
            """
            Non-blocking background thread function for monthly summaries aggregation.
            Properly handles Django database connections in threads.
            """
            try:
                from django.db import connections
                from datetime import date
                from ..utils.utils import run_bulk_aggregation
                from ..models import Tenant
                
                # Close any existing database connections before starting (thread-safe)
                connections.close_all()
                
                # Re-fetch tenant in the thread (ensures fresh DB connection)
                tenant_id = tenant.id
                thread_tenant = Tenant.objects.get(id=tenant_id)
                thread_attendance_date = date(attendance_date.year, attendance_date.month, 1)
                
                logger.info(f"🧵 BACKGROUND THREAD: Starting monthly aggregation for {thread_tenant.subdomain} - {thread_attendance_date.year}-{thread_attendance_date.month:02d}")
                print(f"🧵 BACKGROUND THREAD: Starting monthly aggregation for {thread_tenant.subdomain} - {thread_attendance_date.year}-{thread_attendance_date.month:02d}")
                
                import time
                aggregation_start = time.time()
                result = run_bulk_aggregation(thread_tenant, thread_attendance_date)
                aggregation_elapsed = time.time() - aggregation_start
                
                # Log detailed completion info
                status = result.get('status', 'unknown')
                stats = result.get('statistics', {})
                perf = result.get('performance', {})
                
                logger.info(f"✅ BACKGROUND THREAD: Monthly aggregation completed with status: {status}")
                logger.info(f"📊 BACKGROUND THREAD: Employees processed: {stats.get('employees_processed', 0)}, "
                          f"Records: {stats.get('daily_records_processed', 0)}, "
                          f"Created: {stats.get('attendance_records_created', 0)}, "
                          f"Updated: {stats.get('attendance_records_updated', 0)}")
                logger.info(f"⏱️ BACKGROUND THREAD: Total time: {aggregation_elapsed:.3f}s, "
                          f"DB time: {perf.get('database_time', 'N/A')}, "
                          f"Cache time: {perf.get('cache_clear_time', 'N/A')}")
                
                # Also print to console for visibility
                print(f"✅ BACKGROUND THREAD COMPLETED: {status} - {aggregation_elapsed:.3f}s - "
                      f"{stats.get('employees_processed', 0)} employees, "
                      f"{stats.get('attendance_records_created', 0)} created, "
                      f"{stats.get('attendance_records_updated', 0)} updated")
                
                # Close database connections after thread completes
                connections.close_all()
            except Exception as e:
                logger.error(f"❌ BACKGROUND THREAD ERROR: {str(e)}", exc_info=True)
                # Ensure connections are closed even on error
                try:
                    from django.db import connections
                    connections.close_all()
                except Exception:
                    pass
        
        
        # Start non-blocking background processing thread
        if employee_ids:
            logger.info(f"🧵 ASYNC SUMMARY: About to start background thread for {len(employee_ids)} employees (non-blocking)")
            
            tenant_id = tenant.id if tenant else 'unknown'
            background_thread = threading.Thread(
                target=process_summaries_background, 
                daemon=True,
                name=f"monthly_summary_{tenant_id}_{attendance_date.year}_{attendance_date.month}"
            )
            background_thread.start()
            
            logger.info(f"🧵 ASYNC SUMMARY: Background thread started successfully - Thread ID: {background_thread.ident}, Name: {background_thread.name}")
        else:
            logger.warning(f"⚠️ ASYNC SUMMARY: No employee IDs provided - skipping background processing")
        
        # Return immediately with success response
        total_time = time.time() - start_time
        
        response_data = {
            'message': 'Monthly summary update started',
            'status': 'success'
        }
        
        return Response(response_data, status=200)
        
    except Exception as e:
        logger.error(f"Error in async monthly summary update: {str(e)}")
        return Response({"error": "Failed to start monthly summary update"}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_eligible_employees_for_date(request):
    """
    PROGRESSIVE LOADING API - Load first 50 employees immediately, then lazy load the rest
    
    Two modes:
    1. initial=true: Returns first 50 employees instantly
    2. remaining=true: Returns all remaining employees for lazy loading
    
    PERFORMANCE IMPROVEMENTS:
    - Database-level slicing for instant first batch
    - Bulk attendance lookup with single query
    - Minimized Python loops and processing
    - Optimized data serialization
    
    EXPECTED: ~50ms for first 50, ~200ms for remaining
    """
    try:
        from datetime import datetime, date
        from django.db.models import Q
        from django.core.cache import cache
        from ..models import Attendance
        
        # Performance timing
        start_time = time.time()
        
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return Response({"error": "No tenant found"}, status=400)
        
        date_str = request.query_params.get('date')
        if not date_str:
            return Response({"error": "Date parameter is required"}, status=400)
        
        # PROGRESSIVE LOADING PARAMETERS
        load_initial = request.query_params.get('initial', 'true').lower() == 'true'
        load_remaining = request.query_params.get('remaining', 'false').lower() == 'true'
        
        # Determine batch size and offset based on loading mode
        if load_initial and not load_remaining:
            # Mode 1: Load first 500 employees immediately (increased from 50)
            page_size = 500
            offset = 0
            cache_suffix = 'initial'
            load_mode = 'initial'
        elif load_remaining:
            # Mode 2: Load remaining employees (skip first 500)
            page_size = 2000  # Load all remaining at once
            offset = 500
            cache_suffix = 'remaining'
            load_mode = 'remaining'
        else:
            # Fallback: load first 500 (backward compatibility)
            page_size = 500
            offset = 0
            cache_suffix = 'initial'
            load_mode = 'fallback'
        
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)
        
        # Check if Attendance Excel has been uploaded for this month
        # Excel uploads create Attendance records directly (bulk_create in UploadAttendanceDataAPIView)
        # Signals create Attendance records from DailyAttendance one at a time
        # Key distinction: If Attendance records exist BUT NO DailyAttendance records exist for that month,
        # it means Excel was uploaded (not manual marking via attendance log)
        from ..models import DailyAttendance
        attendance_month_start = date(target_date.year, target_date.month, 1)
        
        # Calculate month end for checking DailyAttendance
        import calendar
        month_end = date(target_date.year, target_date.month, calendar.monthrange(target_date.year, target_date.month)[1])
        
        # Check if Attendance records exist for this month
        attendance_records_exist = Attendance.objects.filter(
            tenant=tenant,
            date=attendance_month_start
        ).exists()
        
        # Check if DailyAttendance records exist for this month (from manual marking)
        daily_records_exist = DailyAttendance.objects.filter(
            tenant=tenant,
            date__gte=attendance_month_start,
            date__lte=month_end
        ).exists()
        
        # Only treat as Excel upload if:
        # 1. Attendance records exist for this month AND
        # 2. NO DailyAttendance records exist for this month (meaning data came from Excel, not manual marking)
        has_excel_attendance = attendance_records_exist and not daily_records_exist
        
        # Check cache first
        cache_key = f"eligible_employees_progressive_{tenant.id}_{date_str}_{cache_suffix}"
        use_cache = request.GET.get('no_cache', '').lower() != 'true'
        
        if use_cache:
            cached_data = cache.get(cache_key)
            if cached_data:
                cached_data['performance'] = {
                    'query_time': f"{(time.time() - start_time):.3f}s",
                    'cached': True,
                    'load_mode': 'initial' if load_initial and not load_remaining else 'remaining'
                }
                # Include Excel attendance flag in cached response
                cached_data['has_excel_attendance'] = has_excel_attendance
                return Response(cached_data)
        
        # Get day of week for off-day checks
        day_of_week = target_date.weekday()
        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        day_name = day_names[day_of_week]
        
        # Build off-day filter efficiently
        off_day_filters = {
            0: Q(off_monday=True),    # Monday
            1: Q(off_tuesday=True),   # Tuesday
            2: Q(off_wednesday=True), # Wednesday
            3: Q(off_thursday=True),  # Thursday
            4: Q(off_friday=True),    # Friday
            5: Q(off_saturday=True),  # Saturday
            6: Q(off_sunday=True),    # Sunday
        }
        off_day_filter = off_day_filters.get(day_of_week, Q())
        
        # PROGRESSIVE LOADING: Get total count once (cached for both requests)
        # NOTE: Include ALL employees (even those with off days) in the count
        total_count_cache_key = f"total_eligible_count_{tenant.id}_{date_str}"
        total_count = cache.get(total_count_cache_key)
        
        if total_count is None:
            total_count = EmployeeProfile.objects.filter(
                tenant=tenant,
                is_active=True
            ).exclude(
                date_of_joining__gt=target_date
            ).count()
            # Cache total count for 5 minutes
            cache.set(total_count_cache_key, total_count, 300)
        
        # OPTIMIZATION 1: Database-level slicing for progressive loading
        # NOTE: Include ALL employees (even those with off days) so they can be shown as "off day"
        eligible_employees_query = EmployeeProfile.objects.filter(
            tenant=tenant,
            is_active=True
        ).exclude(
            date_of_joining__gt=target_date
        ).only(
            # OPTIMIZATION 2: Only fetch required fields (include off day flags)
            'employee_id', 'first_name', 'last_name', 'department',
            'shift_start_time', 'shift_end_time', 'is_active', 'date_of_joining',
            'off_monday', 'off_tuesday', 'off_wednesday', 'off_thursday',
            'off_friday', 'off_saturday', 'off_sunday'
        ).order_by('employee_id')[offset:offset + page_size]  # Progressive loading slice
        
        # OPTIMIZATION 4: Single bulk query for all attendance records  
        employee_ids = [emp.employee_id for emp in eligible_employees_query]
        
        # Bulk fetch attendance records in one query (for the selected date)
        attendance_records = DailyAttendance.objects.filter(
            tenant=tenant,
            date=target_date,
            employee_id__in=employee_ids
        ).only(
            'employee_id', 'attendance_status', 'ot_hours', 'late_minutes', 'check_in', 'check_out'
        )
        
        # OPTIMIZATION 5: Build attendance lookup dictionary efficiently
        attendance_lookup = {
            record.employee_id: {
                'status': record.attendance_status,
                'ot_hours': float(record.ot_hours),
                'late_minutes': record.late_minutes,
                'check_in': record.check_in.strftime('%H:%M') if record.check_in else None,
                'check_out': record.check_out.strftime('%H:%M') if record.check_out else None,
            }
            for record in attendance_records
        }
        
        # Sunday bonus logic: if target date is Sunday and tenant has sunday_bonus_enabled,
        # check previous week's attendance (Mon-Sat) for each employee and mark
        # eligible Sunday as bonus present even if it's an off day.
        sunday_bonus_map = {}
        sunday_bonus_enabled = getattr(tenant, 'sunday_bonus_enabled', False)
        # Present threshold is complement of absent threshold: 7 - weekly_absent_threshold
        weekly_absent_threshold = getattr(tenant, 'weekly_absent_threshold', 4) or 4
        sunday_bonus_threshold = 7 - weekly_absent_threshold
        
        if sunday_bonus_enabled and day_of_week == 6 and employee_ids:
            from datetime import timedelta
            week_start = target_date - timedelta(days=6)  # Previous Monday (for Sun)
            week_end = target_date - timedelta(days=1)    # Up to Saturday
            
            week_attendance_qs = DailyAttendance.objects.filter(
                tenant=tenant,
                employee_id__in=employee_ids,
                date__gte=week_start,
                date__lte=week_end
            ).only('employee_id', 'attendance_status')
            
            # Aggregate present count per employee for the week
            present_counts = {}
            for rec in week_attendance_qs:
                if rec.attendance_status in ['PRESENT', 'PAID_LEAVE']:
                    present_counts[rec.employee_id] = present_counts.get(rec.employee_id, 0) + 1
            
            for emp_id, count in present_counts.items():
                if count >= sunday_bonus_threshold:
                    sunday_bonus_map[emp_id] = True
        
        # Check if employee has off day for the selected date
        def has_off_day_for_date(employee, day_of_week):
            """Check if employee has off day for the given day of week"""
            if day_of_week == 0: return employee.off_monday
            elif day_of_week == 1: return employee.off_tuesday
            elif day_of_week == 2: return employee.off_wednesday
            elif day_of_week == 3: return employee.off_thursday
            elif day_of_week == 4: return employee.off_friday
            elif day_of_week == 5: return employee.off_saturday
            elif day_of_week == 6: return employee.off_sunday
            return False
        
        # OPTIMIZATION 4: Efficient data serialization with minimal processing
        eligible_employees = []
        for employee in eligible_employees_query:
            # Check if this employee has off day for the selected date
            has_off = has_off_day_for_date(employee, day_of_week)
            
            current_attendance = attendance_lookup.get(employee.employee_id, {})
            
            # Quick status determination
            # If employee has off day, set status to 'off' regardless of attendance,
            # UNLESS this is a Sunday bonus day where we intentionally treat them as present.
            is_sunday_bonus = bool(sunday_bonus_map.get(employee.employee_id, False)) if day_of_week == 6 else False
            
            # Also check if attendance record exists and was marked as PRESENT on Sunday (could be from background thread)
            # If it's Sunday and status is PRESENT, check if it meets Sunday bonus criteria
            if day_of_week == 6 and current_attendance and current_attendance['status'] == 'PRESENT':
                # If Sunday bonus is enabled and we haven't already determined it's a bonus,
                # check if the previous week's attendance meets the threshold
                if sunday_bonus_enabled and not is_sunday_bonus:
                    from datetime import timedelta
                    week_start = target_date - timedelta(days=6)  # Previous Monday
                    week_end = target_date - timedelta(days=1)    # Up to Saturday
                    
                    # Count present days in previous week (Mon-Sat)
                    week_attendance = DailyAttendance.objects.filter(
                        tenant=tenant,
                        employee_id=employee.employee_id,
                        date__gte=week_start,
                        date__lte=week_end
                    ).only('attendance_status')
                    
                    present_count = sum(1 for rec in week_attendance 
                                      if rec.attendance_status in ['PRESENT', 'PAID_LEAVE'])
                    
                    # Calculate present threshold (complement of absent threshold)
                    weekly_absent_threshold = getattr(tenant, 'weekly_absent_threshold', 4) or 4
                    present_threshold = 7 - weekly_absent_threshold
                    
                    # If present count meets or exceeds threshold, this Sunday is a bonus
                    if present_count >= present_threshold:
                        is_sunday_bonus = True
                elif not is_sunday_bonus:
                    # If Sunday bonus is not enabled but it's Sunday and marked as PRESENT,
                    # it might have been marked manually, so don't treat as bonus
                    is_sunday_bonus = False
            
            if has_off and is_sunday_bonus:
                # Override OFF to PRESENT for Sunday bonus
                default_status = 'present'
            elif has_off:
                default_status = 'off'
            elif current_attendance:
                default_status = 'present' if current_attendance['status'] in ['PRESENT', 'PAID_LEAVE'] else 'absent'
            else:
                # No attendance record exists - leave unmarked
                default_status = None
            
            # Minimal data processing
            eligible_employees.append({
                'employee_id': employee.employee_id,
                'name': f"{employee.first_name} {employee.last_name}",
                'first_name': employee.first_name,
                'last_name': employee.last_name,
                'department': employee.department or 'General',
                'shift_start_time': employee.shift_start_time.strftime('%H:%M') if employee.shift_start_time else '09:00',
                'shift_end_time': employee.shift_end_time.strftime('%H:%M') if employee.shift_end_time else '18:00',
                'default_status': default_status,
                'has_off_day': has_off,  # Add flag to indicate off day
                'current_attendance': current_attendance,
                'ot_hours': current_attendance.get('ot_hours', 0),
                'late_minutes': current_attendance.get('late_minutes', 0),
                'sunday_bonus': is_sunday_bonus,
            })
        
        # PROGRESSIVE LOADING METADATA
        is_initial_load = load_initial and not load_remaining
        is_remaining_load = load_remaining
        remaining_count = max(0, total_count - 50) if is_initial_load else 0
        
        response_data = {
            'date': date_str,
            'day_name': day_name,
            'eligible_employees': eligible_employees,
            'has_excel_attendance': has_excel_attendance,  # Flag indicating Excel attendance exists for this month
            'progressive_loading': {
                'is_initial_load': is_initial_load,
                'is_remaining_load': is_remaining_load,
                'employees_in_batch': len(eligible_employees),
                'total_employees': total_count,
                'remaining_employees': remaining_count,
                'has_more': remaining_count > 0 if is_initial_load else False,
                'next_batch_url': f"/api/eligible-employees/?date={date_str}&remaining=true" if remaining_count > 0 and is_initial_load else None,
                'preserve_user_changes': True,  # Frontend should preserve user modifications
                'auto_trigger_remaining': is_initial_load and remaining_count > 0,  # Should auto-trigger background load
                'batch_offset': offset,
                'recommended_delay_ms': 100  # Suggested delay before background load
            },
            'total_count': len(eligible_employees),
            'performance': {
                'query_time': f"{(time.time() - start_time):.3f}s",
                'cached': False,
                'load_mode': 'initial' if is_initial_load else 'remaining',
                'batch_size': len(eligible_employees),
                'total_employees': total_count
            }
        }
        
        # Cache for 2 minutes
        if use_cache:
            cache.set(cache_key, response_data, 120)
        
        return Response(response_data)
        
    except Exception as e:
        logger.error(f"Error getting eligible employees: {str(e)}")
        return Response({"error": "Failed to get eligible employees"}, status=500)


class UploadAttendanceDataAPIView(APIView):
    """
    API endpoint for uploading monthly attendance summary data from Excel files
    Only accepts monthly summary format with columns: Employee ID, Name, Department, Present Days, Absent Days, OT Hours, Late Minutes
    """
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        import time
        start_time = time.time()
        
        try:
            import calendar
            import pandas as pd
            from datetime import datetime, date
            from django.db import transaction
            from ..models import Attendance, EmployeeProfile
            
            # Get tenant
            tenant = getattr(request, 'tenant', None)
            if not tenant:
                return Response({
                    'error': 'No tenant found for this request'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get file and parameters
            file_obj = request.FILES.get('file')
            month = request.data.get('month')
            year = request.data.get('year')
            
            if not file_obj:
                return Response({
                    'error': 'No file provided'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if not month or not year:
                return Response({
                    'error': 'Month and year are required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate file type
            if not (file_obj.name.endswith('.xlsx') or file_obj.name.endswith('.xls')):
                return Response({
                    'error': 'Unsupported file format. Please upload Excel (.xlsx, .xls) files only.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                # Read Excel file
                df = pd.read_excel(file_obj)
                
                # Check for monthly summary format (Employee ID and Working Days are optional)
                required_columns = ['Name', 'Department', 'Present Days', 'Absent Days', 'OT Hours', 'Late Minutes']
                optional_columns = ['Working Days']
                
                is_monthly_format = all(col in df.columns for col in required_columns)
                has_working_days = 'Working Days' in df.columns
                has_employee_id = False  # Employee ID no longer expected in template
                
                if not is_monthly_format:
                    return Response({
                        'error': f'Invalid file format. Expected monthly summary columns: {", ".join(required_columns)}',
                        'available_columns': list(df.columns),
                        'required_columns': required_columns,
                        'optional_columns': optional_columns
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Process data
                records_created = 0
                records_updated = 0
                errors = []
                warnings = []
                
                # Process monthly summary format only
                # Auto-generate Employee IDs for entries that don't have them
                from ..utils.utils import generate_employee_id_bulk_optimized
                
                # OPTIMIZED: Use values() to reduce data transfer, iterator for memory efficiency
                existing_employees_by_name = {}
                for emp in EmployeeProfile.objects.filter(
                    tenant=tenant, is_active=True
                ).values('first_name', 'last_name', 'employee_id').iterator(chunk_size=1000):
                    key = f"{emp['first_name']} {emp['last_name']}".strip().lower()
                    existing_employees_by_name[key] = emp['employee_id']
                
                # Prepare data for Employee ID generation and matching
                employees_data = []
                final_employee_ids = []
                
                for index, row in df.iterrows():
                    name = str(row['Name']).strip()
                    department = str(row.get('Department', '')).strip()
                    
                    # Employee ID will be auto-generated based on name matching
                    provided_id = ''
                    
                    if provided_id:
                        # Use provided Employee ID
                        final_employee_ids.append(provided_id)
                    else:
                        # Try to match by name first
                        name_key = name.lower()
                        if name_key in existing_employees_by_name:
                            # Found existing employee by name
                            final_employee_ids.append(existing_employees_by_name[name_key])
                        else:
                            # Need to generate new Employee ID
                            employees_data.append({
                                'name': name,
                                'department': department
                            })
                            final_employee_ids.append(None)  # Will be filled after generation
                
                # Generate Employee IDs for new employees only
                if employees_data:
                    employee_id_mapping = generate_employee_id_bulk_optimized(employees_data, tenant.id)
                    
                    # Fill in the generated IDs
                    generated_index = 0
                    for i, emp_id in enumerate(final_employee_ids):
                        if emp_id is None:
                            final_employee_ids[i] = employee_id_mapping[generated_index]
                            generated_index += 1
                
                # Add final Employee IDs to the dataframe
                df['Final_Employee_ID'] = final_employee_ids
                
                # OPTIMIZED: Reuse existing_employees_by_name to avoid second query
                employee_ids = df['Final_Employee_ID'].dropna().unique()
                existing_employee_set = set(existing_employees_by_name.values())
                missing_employees = set(employee_ids) - existing_employee_set
                
                # If there are missing employees, collect their details and return for confirmation
                if missing_employees:
                    missing_employee_details = []
                    for index, row in df.iterrows():
                        employee_id = str(row['Final_Employee_ID']).strip()
                        if employee_id in missing_employees:
                            name = str(row['Name']).strip()
                            department = str(row.get('Department', '')).strip()
                            
                            # Split name into first and last name
                            name_parts = name.split(' ', 1)
                            first_name = name_parts[0] if name_parts else ''
                            last_name = name_parts[1] if len(name_parts) > 1 else ''
                            
                            # Clean up last name - avoid "nan" values
                            if not last_name or str(last_name).lower() in ['nan', 'none', '']:
                                last_name = ''
                            
                            missing_employee_details.append({
                                'employee_id': employee_id,
                                'name': name,
                                'first_name': first_name,
                                'last_name': last_name,
                                'department': department,
                                'row_number': index + 2  # Excel row number (accounting for header)
                            })
                    
                    return Response({
                        'error': 'Missing employees found',
                        'missing_employees': missing_employee_details,
                        'total_missing': len(missing_employee_details),
                        'message': f'Found {len(missing_employee_details)} employees that do not exist in the system. Please confirm to create them.'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # ULTRA-OPTIMIZED: Use bulk operations instead of N queries
                from datetime import date
                attendance_date = date(int(year), int(month), 1)
                calendar_days = calendar.monthrange(int(year), int(month))[1]
                
                # Get existing attendance records for this period (single query)
                existing_attendance_dict = {
                    att['employee_id']: att['id']
                    for att in Attendance.objects.filter(
                        tenant=tenant,
                        date=attendance_date
                    ).values('employee_id', 'id')
                }
                
                # Prepare bulk create and update lists
                attendance_to_create = []
                attendance_to_update = []
                
                for index, row in df.iterrows():
                    try:
                        employee_id = str(row['Final_Employee_ID']).strip()
                        
                        # Process monthly summary format
                        name = str(row['Name']).strip()
                        department = str(row.get('Department', '')).strip()
                        
                        # Validate employee exists
                        if employee_id not in existing_employee_set:
                            errors.append(f'Row {index + 2}: Employee {employee_id} not found or inactive')
                            continue
                        
                        # Get numeric values
                        present_days = float(row.get('Present Days', 0)) if pd.notna(row.get('Present Days')) else 0
                        absent_days = float(row.get('Absent Days', 0)) if pd.notna(row.get('Absent Days')) else 0
                        ot_hours = float(row.get('OT Hours', 0)) if pd.notna(row.get('OT Hours')) else 0
                        late_minutes = int(row.get('Late Minutes', 0)) if pd.notna(row.get('Late Minutes')) else 0
                        
                        # IMPORTANT: absent_days from Excel is preserved as-is
                        # The model's save() method only auto-calculates when absent_days=0
                        # This ensures Excel values are respected during bulk upload
                        
                        # OPTIMIZED: Use standard 30 days for bulk performance (DOJ calc in background if needed)
                        if has_working_days and pd.notna(row.get('Working Days')):
                            total_working_days = float(row.get('Working Days', 0))
                            if total_working_days <= 0:
                                total_working_days = 30
                        else:
                            total_working_days = 30
                        
                        # For Excel uploads, unmarked_days should always be 0
                        # because uploading Excel means the entire month is being marked/processed
                        # This ensures that the calendar days for this month are excluded from total unmarked days
                        unmarked_days = 0
                        
                        attendance_data = {
                            'tenant': tenant,
                            'employee_id': employee_id,
                            'name': name,
                            'department': department,
                            'date': attendance_date,
                            'calendar_days': calendar_days,
                            'total_working_days': int(total_working_days),
                            'present_days': present_days,
                            'absent_days': absent_days,
                            'unmarked_days': unmarked_days,
                            'ot_hours': ot_hours,
                            'late_minutes': late_minutes
                        }
                        
                        if employee_id in existing_attendance_dict:
                            # Update existing record
                            attendance_data['id'] = existing_attendance_dict[employee_id]
                            attendance_to_update.append(Attendance(**attendance_data))
                            records_updated += 1
                        else:
                            # Create new record
                            attendance_to_create.append(Attendance(**attendance_data))
                            records_created += 1
                            
                    except Exception as e:
                        errors.append(f'Row {index + 2}: {str(e)}')
                
                # BULK OPERATIONS: 10-100x faster than individual saves!
                with transaction.atomic():
                    if attendance_to_create:
                        Attendance.objects.bulk_create(attendance_to_create, batch_size=100, ignore_conflicts=True)
                    
                    if attendance_to_update:
                        Attendance.objects.bulk_update(
                            attendance_to_update,
                            ['name', 'department', 'total_working_days', 'present_days', 
                             'absent_days', 'unmarked_days', 'ot_hours', 'late_minutes', 'calendar_days'],
                            batch_size=100
                        )
                
                
                # Clear relevant caches
                from django.core.cache import cache
                cache_keys = [
                    f"payroll_overview_{tenant.id}",
                    f"attendance_all_records_{tenant.id}",
                    f"directory_data_{tenant.id}",
                    f"directory_data_full_{tenant.id}",  # Clear full directory cache
                    f"months_with_attendance_{tenant.id}"
                ]
                for key in cache_keys:
                    cache.delete(key)
                
                # Clear frontend charts cache (stats component)
                try:
                    cache.delete_pattern(f"frontend_charts_{tenant.id}_*")
                    logger.info(f"✨ Cleared frontend_charts pattern cache for tenant {tenant.id}")
                except AttributeError:
                    # Fallback: Clear common chart cache keys
                    chart_keys = [
                        f"frontend_charts_{tenant.id}_this_month_All_",
                        f"frontend_charts_{tenant.id}_last_6_months_All_",
                        f"frontend_charts_{tenant.id}_last_12_months_All_",
                        f"frontend_charts_{tenant.id}_last_5_years_All_"
                    ]
                    for key in chart_keys:
                        cache.delete(key)
                    logger.info(f"✨ Cleared frontend_charts fallback cache for tenant {tenant.id}")
                
                logger.info(f"✨ Cleared directory and charts cache for tenant {tenant.id} after attendance upload")
                
                # Calculate upload time
                upload_time = round((time.time() - start_time) * 1000, 2)  # milliseconds
                
                response_data = {
                    'message': 'Attendance data uploaded successfully',
                    'records_created': records_created,
                    'records_updated': records_updated,
                    'month': month,
                    'year': year
                }
                
                if errors:
                    response_data['total_errors'] = len(errors)
                    response_data['errors'] = errors[:10]  # Show first 10 errors
                    
                if warnings:
                    response_data['total_warnings'] = len(warnings)
                    response_data['warnings'] = warnings[:10]  # Show first 10 warnings
                
                return Response(response_data, status=status.HTTP_201_CREATED)
                
            except Exception as e:
                # SECURITY: Don't expose internal error details to client
                logger.error(f"Error in attendance upload: {str(e)}", exc_info=True)
                return Response({
                    'error': 'Failed to process file. Please try again.'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
        except Exception as e:
            logger.error(f"Error in attendance upload: {str(e)}")
            return Response({
                'error': 'Upload failed. Please try again.'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UploadMonthlyAttendanceAPIView(APIView):
    """
    API endpoint for uploading monthly attendance summary data
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            import calendar
            from datetime import date
            from django.db import transaction
            from ..models import Attendance
            
            # Get tenant
            tenant = getattr(request, 'tenant', None)
            if not tenant:
                return Response({
                    'error': 'No tenant found for this request'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get parameters
            month = request.data.get('month')
            year = request.data.get('year')
            data = request.data.get('data', [])
            
            if not month or not year:
                return Response({
                    'error': 'Month and year are required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if not data:
                return Response({
                    'error': 'No attendance data provided'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            created = 0
            failed = 0
            errors = []
            
            # Create attendance records
            attendance_records = []
            
            for index, record in enumerate(data):
                try:
                    employee_id = record.get('employee_id')
                    name = record.get('name')
                    
                    if not employee_id or not name:
                        errors.append(f'Record {index + 1}: Missing employee_id or name')
                        failed += 1
                        continue
                    
                    # Create attendance date (first day of the month)
                    attendance_date = date(int(year), int(month), 1)
                    
                    # For Excel/JSON uploads, unmarked_days should always be 0
                    # because uploading means the entire month is being marked/processed
                    unmarked_days = 0
                    
                    attendance_record = Attendance(
                        tenant=tenant,
                        employee_id=employee_id,
                        name=name,
                        department=record.get('department', ''),
                        date=attendance_date,
                        calendar_days=calendar.monthrange(int(year), int(month))[1],
                        total_working_days=record.get('total_working_days', 0),
                        present_days=record.get('present_days', 0),
                        absent_days=record.get('absent_days', 0),
                        unmarked_days=unmarked_days,
                        ot_hours=record.get('ot_hours', 0),
                        late_minutes=record.get('late_minutes', 0)
                    )
                    
                    attendance_records.append(attendance_record)
                    created += 1
                    
                except Exception as e:
                    errors.append(f'Record {index + 1}: {str(e)}')
                    failed += 1
            
            # Bulk create records
            if attendance_records:
                with transaction.atomic():
                    Attendance.objects.bulk_create(attendance_records, ignore_conflicts=True)
            
            # Clear directory cache after successful upload
            from django.core.cache import cache
            cache_keys = [
                f"directory_data_{tenant.id}",
                f"directory_data_full_{tenant.id}",  # Clear full directory cache
                f"attendance_all_records_{tenant.id}",
                f"months_with_attendance_{tenant.id}",  # Clear months-with-attendance endpoint cache
                f"payroll_overview_{tenant.id}"  # Clear payroll overview cache
            ]
            for key in cache_keys:
                cache.delete(key)
            
            # Clear frontend charts cache (stats component)
            try:
                cache.delete_pattern(f"frontend_charts_{tenant.id}_*")
                logger.info(f"✨ Cleared frontend_charts pattern cache for tenant {tenant.id}")
            except AttributeError:
                # Fallback: Clear common chart cache keys
                chart_keys = [
                    f"frontend_charts_{tenant.id}_this_month_All_",
                    f"frontend_charts_{tenant.id}_last_6_months_All_",
                    f"frontend_charts_{tenant.id}_last_12_months_All_",
                    f"frontend_charts_{tenant.id}_last_5_years_All_"
                ]
                for key in chart_keys:
                    cache.delete(key)
                logger.info(f"✨ Cleared frontend_charts fallback cache for tenant {tenant.id}")
            
            logger.info(f"✨ Cleared directory and charts cache for tenant {tenant.id} after monthly attendance upload")
            
            return Response({
                'message': 'Monthly attendance data uploaded successfully',
                'total_records': len(data),
                'created': created,
                'failed': failed,
                'errors': errors[:10],  # Show first 10 errors
                'month': month,
                'year': year
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f"Error in monthly attendance upload: {str(e)}")
            return Response({
                'error': 'Upload failed. Please try again.'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DownloadAttendanceTemplateAPIView(APIView):
    """
    API endpoint for downloading attendance template
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from openpyxl import Workbook
            from django.http import HttpResponse
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Template"
            
            # Define headers for monthly attendance format (Employee ID is optional)
            headers = [
                'Name', 'Department', 'Present Days', 'Absent Days', 'Working Days', 'OT Hours', 'Late Minutes'
            ]
            
            # Add headers to worksheet
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add sample data for monthly format (Employee IDs will be auto-generated based on name matching)
            sample_data = [
                ['John Doe', 'Engineering', 22, 3, 25, 15.5, 45],
                ['Jane Smith', 'HR', 20, 5, 25, 8.0, 30],
                ['Bob Johnson', 'Sales', 18, 7, 25, 12.0, 60],
                ['Alice Brown', 'Marketing', 21, 4, 25, 5.5, 25],
                ['Charlie Wilson', 'Finance', 19, 6, 25, 10.0, 40]
            ]
            
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Add instructions
            ws.cell(row=8, column=1, value="Instructions:")
            ws.cell(row=9, column=1, value="1. Name: Full name of the employee (Employee ID will be auto-generated)")
            ws.cell(row=10, column=1, value="2. Department: Employee's department")
            ws.cell(row=11, column=1, value="3. Present Days: Number of days employee was present")
            ws.cell(row=12, column=1, value="4. Absent Days: Number of days employee was absent")
            ws.cell(row=13, column=1, value="5. Working Days: Total working days (optional - will calculate from DOJ if not provided)")
            ws.cell(row=14, column=1, value="6. OT Hours: Total overtime hours (decimal, e.g., 15.5)")
            ws.cell(row=15, column=1, value="7. Late Minutes: Total late minutes (integer, e.g., 45)")
            ws.cell(row=16, column=1, value="8. This is a MONTHLY SUMMARY format - no individual dates needed")
            
            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=attendance_template.xlsx'
            wb.save(response)
            
            return response
            
        except Exception as e:
            return Response({
                'error': f'Failed to generate template: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class RevertPenaltyDayView(APIView):
    """
    API endpoint to revert penalty days by marking an employee as PRESENT on a penalty day.
    When an employee is marked PRESENT, the weekly penalty is recalculated automatically.
    If the absent count drops below the threshold, the penalty is removed.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """
        Revert penalty day for an employee.
        
        Request body:
        {
            "employee_id": "EMP001",
            "date": "2025-01-15"
        }
        """
        try:
            employee_id = request.data.get('employee_id')
            date_str = request.data.get('date')

            # Validate input
            if not employee_id or not date_str:
                return Response({
                    'error': 'employee_id and date are required'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Parse date
            from datetime import datetime, timedelta
            try:
                attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({
                    'error': 'Invalid date format. Use YYYY-MM-DD'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Get tenant from request
            tenant = request.user.tenant if hasattr(request.user, 'tenant') else None
            if not tenant:
                return Response({
                    'error': 'Tenant not found'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Get week boundaries (Monday to Sunday)
            day_of_week = attendance_date.weekday()  # 0 = Monday, 6 = Sunday
            week_start = attendance_date - timedelta(days=day_of_week)
            week_end = week_start + timedelta(days=6)

            # TOGGLE: Flip DailyAttendance.penalty_ignored for this specific date
            try:
                daily = DailyAttendance.objects.get(
                    tenant=tenant,
                    employee_id=employee_id,
                    date=attendance_date
                )
            except DailyAttendance.DoesNotExist:
                return Response({
                    'success': False,
                    'error': 'Daily attendance not found for this employee and date'
                }, status=status.HTTP_404_NOT_FOUND)

            old_ignored = bool(getattr(daily, 'penalty_ignored', False))
            new_ignored = not old_ignored
            setattr(daily, 'penalty_ignored', new_ignored)
            daily.save(update_fields=['penalty_ignored'])

            # Compute absent count for the week excluding ignored days (for context)
            from decimal import Decimal
            absent_count = DailyAttendance.objects.filter(
                tenant=tenant,
                employee_id=employee_id,
                date__gte=week_start,
                date__lte=week_end,
                attendance_status='ABSENT',
                penalty_ignored=False
            ).count()

            logger.info(
                f"🔁 Toggled penalty_ignored for {employee_id} on {date_str}: {old_ignored} → {new_ignored}; "
                f"week {week_start.isoformat()}–{week_end.isoformat()} absent(excluding ignored)={absent_count}"
            )

            return Response({
                'success': True,
                'message': 'Penalty reverted (ignored for weekly count)' if new_ignored else 'Penalty un-reverted (counted again)',
                'data': {
                    'employee_id': employee_id,
                    'date': date_str,
                    'penalty_ignored': new_ignored,
                    'week_info': {
                        'week_start': week_start.isoformat(),
                        'week_end': week_end.isoformat(),
                        'absent_count_excluding_ignored': absent_count
                    }
                }
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f'❌ Error reverting penalty: {str(e)}', exc_info=True)
            return Response({
                'error': f'Failed to revert penalty: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
